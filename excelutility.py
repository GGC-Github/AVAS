#!/usr/bin/env python3
import io
import os
from openpyxl import load_workbook
import excelstyle
import analysisutility as utility
import datetime


def makeExcelReport(analysisRes, sysList):
    dt = datetime.datetime.now()

    wb = load_workbook(io.BytesIO(excelstyle.defaultexcel))
    wsResSum = wb['진단 결과 요약']
    impdict = {}
    totalcnt = 0
    resultcnt = 0

#   자산 정보
    wsResSum['D6'] = sysList['osType']
    wsResSum['D7'] = "{} {}".format(sysList['osName'], sysList['osVersion'])
    wsResSum['D8'] = sysList['hostname']
    wsResSum['D9'] = dt.strftime("%Y-%m-%d %H:%M:%S")

#   진단 현황 분포표
    impdict.update({'양호': [int(data[3]['ImportantScore']) for data in analysisRes if data[1] in '양호']})
    impdict.update({'취약': [int(data[3]['ImportantScore']) for data in analysisRes if data[1] in '취약']})
    impdict.update({'리뷰': [int(data[3]['ImportantScore']) for data in analysisRes if data[1] in '리뷰']})

    # 상태 별 결과 분포표
    for cnt, val in zip(range(0, 4), [len(analysisRes), len(impdict['양호']), len(impdict['취약']), len(impdict['리뷰'])]):
        wsResSum.cell(row=19, column=10 + cnt).font = excelstyle.normalfont
        wsResSum.cell(row=19, column=10 + cnt).value = val

    # 중요도 별 결과 분포표
    for colcnt, chkval in zip(range(0, 3), ['양호', '취약', '리뷰']):
        for rowcnt in range(0, 3):
            valcnt = impdict[chkval].count(rowcnt + 1)
            wsResSum.cell(row=29 + rowcnt, column=11 + colcnt).font = excelstyle.normalfont
            wsResSum.cell(row=29 + rowcnt, column=11 + colcnt).value = valcnt
            totalcnt += wsResSum.cell(row=29 + rowcnt, column=10).value * valcnt
            if colcnt == 0:
                resultcnt += wsResSum.cell(row=29 + rowcnt, column=10).value * valcnt
            elif colcnt == 2:
                resultcnt += wsResSum.cell(row=29 + rowcnt, column=10).value * valcnt * 0.5

    # 보안 점수 결과
    wsResSum['D2'] = resultcnt / totalcnt * 100
    wsResSum['D2'].font = excelstyle.headredfont

#   진단 결과 내역
    analysisRes.sort(key=lambda x: x[3]['Category'])
    for cnt in range(0, len(analysisRes)):
        inputdata = [analysisRes[cnt][3]['Category'], analysisRes[cnt][0], analysisRes[cnt][3]['Name'],
                     int(analysisRes[cnt][3]['ImportantScore']), analysisRes[cnt][1]]
        for num, idx in zip([0, 2, 3, 10, 11], range(0, len(inputdata))):
            wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.normalfont
            if num == 10 or num == 11:
                wsResSum.cell(row=36 + cnt, column=2 + num).alignment = excelstyle.centeralign

            if num == 11:
                if inputdata[idx] == '양호':
                    wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.greenfont
                elif inputdata[idx] == '리뷰':
                    wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.bluefont
                elif inputdata[idx] == '취약':
                    wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.redfont

            wsResSum.cell(row=36 + cnt, column=2 + num).border = excelstyle.thinborder
            wsResSum.cell(row=36 + cnt, column=2 + num).value = inputdata[idx]
            if num == 3:
                wsResSum.merge_cells(start_row=36 + cnt, start_column=2 + num, end_row=36 + cnt, end_column=2 + num + 6)
            elif num == 0:
                wsResSum.merge_cells(start_row=36 + cnt, start_column=2 + num, end_row=36 + cnt, end_column=2 + num + 1)

#   진단 결과 상세 내역 (세로)
    wsResDetVer = wb['진단 결과 상세']
    rownum = 3
    for cnt in range(0, len(analysisRes)):
        inputdata = [
            [' ', ' '], ['구분', analysisRes[cnt][3]['Category']], ['코드', analysisRes[cnt][0]],
            ['항목', analysisRes[cnt][3]['Name']], ['중요도', int(analysisRes[cnt][3]['ImportantScore'])],
            ['진단 결과', analysisRes[cnt][1]], ['판단 기준', analysisRes[cnt][3]['Criterion']],
            ['상세 현황', analysisRes[cnt][2]], ['조치 방법', analysisRes[cnt][3]['ActionPlan']]
        ]
        for idx in range(0, 9):
            for colcnt in range(0, 2):
                if idx != 0:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).border = excelstyle.thinborder
                if colcnt == 0 and idx != 0:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).alignment = excelstyle.centerwrapalign
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).fill = excelstyle.cellbgfill
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).font = excelstyle.whiteboldfont
                else:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).alignment = excelstyle.leftwrapalign
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).font = excelstyle.normalfont

                if idx == 7 and colcnt == 1:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).value = utility.mergeExeclData(inputdata[idx][colcnt])
                else:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).value = inputdata[idx][colcnt]

            rownum += 1

#   진단 결과 상세 내역 (가로)
    wsResDetHor = wb['진단 결과 상세(가로)']
    for cnt in range(0, len(analysisRes)):
        inputdata = [
            analysisRes[cnt][3]['Category'], analysisRes[cnt][0], analysisRes[cnt][3]['Name'], int(analysisRes[cnt][3]['ImportantScore']),
            analysisRes[cnt][1], analysisRes[cnt][3]['Criterion'], analysisRes[cnt][2], analysisRes[cnt][3]['ActionPlan']
        ]
        for idx in range(0, 8):
            wsResDetHor.cell(row=5 + cnt, column=2 + idx).border = excelstyle.thinborder
            wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.normalfont
            if idx == 6:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).value = utility.mergeExeclData(inputdata[idx])
            else:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).value = inputdata[idx]
            if idx < 2:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).alignment = excelstyle.leftalign
            elif idx == 4 or idx == 3:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).alignment = excelstyle.centeralign
                if inputdata[idx] == '양호':
                    wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.greenfont
                elif inputdata[idx] == '리뷰':
                    wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.bluefont
                elif inputdata[idx] == '취약':
                    wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.redfont
            else:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).alignment = excelstyle.fillalign

    wsResRefInfo = wb['참고']
    for info, val in zip(['ipList', 'processInfo', 'portInfo', 'serviceInfo'], ['B5', 'B31', 'B57', 'B83']):
        if info in sysList.keys():
            wsResRefInfo[val] = sysList[info]
            wsResRefInfo[val].font = excelstyle.normalfont
            wsResRefInfo[val].alignment = excelstyle.leftwrapalign

    fileName = ''.join("result_report_{}_{}_{}.xlsx".format(
        sysList['osName'], sysList['hostname'], dt.strftime("%Y%m%d%H%M%S"))
    )

    wb.save(filename=os.path.join(os.getcwd(), 'ExcelDir', fileName))
